import pandas as pd
import openpyxl

def check_for_hyperlinks():
    # Load workbook with openpyxl to check for hyperlinks
    wb = openpyxl.load_workbook("translations_spreadsheet.xlsx")
    
    for sheet_name in wb.sheetnames:
        print(f"=== {sheet_name} ===")
        ws = wb[sheet_name]
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.hyperlink:
                    print(f"Hyperlink found in {cell.coordinate}: {cell.hyperlink.target}")
                    print(f"Cell value: {cell.value}")
        print()

if __name__ == "__main__":
    check_for_hyperlinks()